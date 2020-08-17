from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import base64
import json


class Results():
    def __init__(self):
        self.differentiator = []
        self.captures = [] 
        self.signals = []
        self.captures_images = []
        self.differentiator_image = None
    

    def initialize_parameters(self, total_time, captures_seg, description):
        self.initial_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.total_time = total_time
        self.captures_seg = captures_seg
        self.interval = (1 / self.captures_seg)
        self.description = description or ""
        self.captures.clear()
        self.signals.clear()
        self.captures_images.clear()
    

    def get_differentiator_image(self, webcam):
        try:
            return self.encode_image(self.differentiator_image, webcam)
        except:
            return ''


    def get_all_images(self, webcam):
        try:
            return self.encode_all_images(webcam)
        except:
            return ''
    

    def encode_all_images(self, webcam):
        encoded = {}
        encoded["diferenciador.jpg"] = f"{self.encode_image(self.differentiator_image, webcam)}"
        for i in range(0, len(self.captures_images)):
            image = self.captures_images[i]
            encoded[f"captura_{i + 1}.jpg"] = f"{self.encode_image(image, webcam)}"
        return json.dumps(encoded)


    def encode_image(self, image, webcam):
        jpg_image = webcam.encode_to_jpg(image)
        return base64.b64encode(jpg_image)
    

    def get_xlsx_results(self):
        try:
            return self.generate_xlsx()
        except:
            return ''
    
    def generate_xlsx(self):
        excel_file = Workbook()
        spreadsheet1 = excel_file.active
        spreadsheet1.title = "Resultados"

        row_number = self.save_basic_information(spreadsheet1)
        row_number = self.generate_differentiator_table(spreadsheet1, row_number + 1)
        self.generate_captures_table(spreadsheet1, row_number + 1)
        self.configure_columns_width(spreadsheet1)
        return self.encode_excel(excel_file)

    def save_basic_information(self, spreadsheet):
        font1 = Font(name='Arial', size=12)
        general_information = self.get_general_information()
        for index, information in enumerate(general_information):
            for position, column in enumerate(information):
                spreadsheet.cell(row = index + 1, column = position + 1).font = font1
                spreadsheet.cell(row = index + 1, column = position + 1, value = column)
        return len(general_information)
    

    def get_general_information(self):
        init = ["Data", self.initial_date]
        duration = ["Duração", f"{self.total_time} segundos"]
        captures = ["Capturas", f"{self.captures_seg} cap./seg."]
        total = ["Total", f"{self.total_time * self.captures_seg} cap."]
        descript = ["Descrição", f"{self.description or 'Não informada'}"]
        return [init, duration, captures, total, descript]
    

    def generate_differentiator_table(self, spreadsheet, row_number):
        font1 = Font(name='Arial', size=12, bold=True)
        font2 = Font(name='Arial', size=12)
        differentiator_information = self.get_differentiator_information()
        for index, information in enumerate(differentiator_information):
            for position, column in enumerate(information):
                spreadsheet.cell(row = row_number + index + 1, column = position + 1).font = font1 if index == 0 else font2
                spreadsheet.cell(row = row_number + index + 1, column = position + 1, value = column)
        return row_number + len(differentiator_information)
    

    def get_differentiator_information(self):
        title = ["Diferenciador"]
        table_headers = ["Vermelho", "Verde", "Azul"]
        table_content = [self.differentiator[2], self.differentiator[1], self.differentiator[0]]
        return [title, table_headers, table_content]
    

    def generate_captures_table(self, spreadsheet, row_number):
        font1 = Font(name='Arial', size=12, bold=True)
        font2 = Font(name='Arial', size=12)
        captures_information = self.get_captures_information()
        for index, information in enumerate(captures_information):
            for position, column in enumerate(information):
                spreadsheet.cell(row = row_number + index + 1, column = position + 1).font = font1 if index == 0 else font2
                spreadsheet.cell(row = row_number + index + 1, column = position + 1, value = column)
    

    def get_captures_information(self):
        title = ["Capturas"]
        table_headers = ["Tempo(seg)", "Vermelho", "Verde", "Azul", "Sinal"]
        information = [title, table_headers]
        for index, value in enumerate(self.captures):
            new_data = [(index + 1) * self.interval, value[2], value[1], value[0], self.signals[index]]
            information.append(new_data)
        return information
    

    def configure_columns_width(self, spreadsheet):
        for column in spreadsheet.columns:
            # length = max(len(str(cell.value)) for cell in column)
            spreadsheet.column_dimensions[column[0].column_letter].width = 25.0 #length * 1.2
    

    def encode_excel(self, excel_file):
        encoded = BytesIO()
        excel_file.save(encoded)
        encoded.seek(0)
        base_64 = base64.b64encode(encoded.read())
        return base_64

