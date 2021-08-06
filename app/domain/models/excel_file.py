from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from app.domain.packs.image_pack import ImagePack


class ExcelFile():
    """ Classe que encapsula toda a lógica de criação e personalização do arquivo xlsx.

    Tambem encapsula os usos da biblioteca openpyxl.
    O arquivo gerado é convertido em bytes na BASE_64.
    """

    def __init__(self, title=''):
        self.excel_file = Workbook()  # Arquivo excel criado com auxílio da bilioteca openpyxl.
        self.spreadsheet = self.excel_file.active  # Criação de um spreadsheet (página de um excel).
        # definição do títlo do spreadsheet (recebido por parâmetro na instanciação da classe).
        self.spreadsheet.title = title
        self.to_merge_cells = []  # Lista de células que irão compor o spreadsheet.

    def create(self, general_info, differentiator_info, captures_info, calibration_info):
        """ Método responsável por conduzir a personalização do arquivo de acordo com as informações recebidas. 

        Os parâmetros são listas de listas contendo títulos e dados.
        Retorna um arquivo xlsx codificado em bytes na base_64.
        """
        row_count = self.generate_table(general_info, 1)
        row_count = self.generate_table(differentiator_info, row_count)
        row_count = self.generate_table(captures_info, row_count)
        calibration_count = self.generate_table(calibration_info, row_count)
        self.configure_columns_width_and_align()
        self.merge_cells()
        self.generate_captures_graph(captures_info, row_count)
        self.generate_calibration_graph(calibration_info, calibration_count)
        return self.encode_excel(self.excel_file)

    def generate_table(self, current_info, row_count):
        """ Método responsável por receber e escrever as informações no spreeedshet.

        Recebe como parâmetro as informações que devem ser escritas e o nº da primeira linha que deve ser usada.
        Retorna um inteiro com a próxima linha válida para continuar registrando informações.
        """
        if len(current_info) == 0:
            return row_count
        font1 = Font(name='Arial', size=14, bold=True, )
        font2 = Font(name='Arial', size=12)
        for i, row in enumerate(current_info):
            for j, column in enumerate(row):
                cell = self.spreadsheet.cell(row=row_count + i, column=j + 1)
                cell.font = font1 if i == 0 else font2
                cell.alignment = Alignment(horizontal='center') if i == 0 or i == 1 else Alignment(horizontal='right')
                cell.value = column
                if i == 0:
                    self.to_merge_cells.append([row_count, len(current_info[0])])
        return row_count + len(current_info) + 2  # 2 linhas em branco

    def configure_columns_width_and_align(self):
        """ Configura o tamanho mínimo das colunas e o alinhamento para que infomações não sejam ocultadas. 

        Os testes feitos apontaram 25 como um bom tamanho.
        Nenhum valor é retornado.
        """
        for column in self.spreadsheet.columns:
            self.spreadsheet.column_dimensions[column[0].column_letter].width = 25.0
        for row in self.spreadsheet.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True)

    def merge_cells(self):
        """ Mescla as colunas de título do spreedshet. """
        for value in self.to_merge_cells:
            self.spreadsheet.merge_cells(start_row=value[0], start_column=1, end_row=value[0], end_column=value[1])
            self.spreadsheet.cell(row=value[0], column=1).fill = PatternFill(fgColor='D3D3D3', fill_type='solid')

    def generate_captures_graph(self, captures, rows):
        """ Gera o gráfico Sinais X Tempo. """
        my_chart = ScatterChart()
        my_chart.title = 'Gráfico dos Sinais'
        my_chart.style = 16
        my_chart.y_axis.title = 'Sinal'
        my_chart.x_axis.title = 'Tempo (segundos)'
        x_values = Reference(self.spreadsheet, min_col=2, min_row=rows - len(captures), max_row=rows - 3)
        y_values = Reference(self.spreadsheet, min_col=6, min_row=rows - len(captures) - 1, max_row=rows - 3)
        series = Series(y_values, x_values, title_from_data=True)
        my_chart.series.append(series)
        my_chart.width = 23
        my_chart.height = 10
        self.spreadsheet.add_chart(my_chart, f"I1")

    def generate_calibration_graph(self, values, rows):
        """ Gera o gráfico sinal X ciclo. """
        if len(values) == 0:
            return
        my_chart = ScatterChart()
        my_chart.title = 'Gráfico de calibração'
        my_chart.style = 16
        my_chart.y_axis.title = 'Sinal'
        my_chart.x_axis.title = 'Ciclo'
        x_values = Reference(self.spreadsheet, min_col=1, min_row=rows - len(values), max_row=rows - 3)
        y_values = Reference(self.spreadsheet, min_col=2, min_row=rows - len(values) - 1, max_row=rows - 3)
        series = Series(y_values, x_values, title_from_data=True)
        series.marker = Marker('circle', 7)
        series.graphicalProperties.line.noFill = True
        my_chart.series.append(series)
        my_chart.width = 23
        my_chart.height = 10
        self.spreadsheet.add_chart(my_chart, f"I22")

    def encode_excel(self, excel_file):
        """ Codifica o arquivo xlsx em bytes na base_64. """
        encoded = ImagePack.create_bytes_file()
        excel_file.save(encoded)
        encoded.seek(0)
        base_64_result = ImagePack.encode_to_b64(encoded.read())
        return base_64_result
